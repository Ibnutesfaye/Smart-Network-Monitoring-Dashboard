import csv
import os
from datetime import timedelta

from django.conf import settings
from django.db import models
from django.utils import timezone
from openpyxl import Workbook

from apps.alerts.models import Alert
from apps.audit.models import ActivityLog
from apps.devices.models import Device
from apps.traffic.models import TrafficSample


def _period_bounds(report_type):
    now = timezone.now()
    if report_type == "daily":
        start = now - timedelta(days=1)
    elif report_type == "weekly":
        start = now - timedelta(days=7)
    else:
        start = now - timedelta(days=30)
    return start, now


def collect_report_data(period_start, period_end):
    devices = Device.objects.all()
    online = devices.filter(status=Device.Status.ONLINE).count()
    offline = devices.filter(status=Device.Status.OFFLINE).count()
    traffic = TrafficSample.objects.filter(timestamp__range=(period_start, period_end))
    alerts = Alert.objects.filter(created_at__range=(period_start, period_end))
    logs = ActivityLog.objects.filter(created_at__range=(period_start, period_end))
    avg_bw = traffic.aggregate(avg=models.Avg("bandwidth_usage"))["avg"] or 0
    return {
        "period_start": str(period_start),
        "period_end": str(period_end),
        "total_devices": devices.count(),
        "online_devices": online,
        "offline_devices": offline,
        "traffic_samples": traffic.count(),
        "avg_bandwidth": round(avg_bw, 2),
        "total_alerts": alerts.count(),
        "critical_alerts": alerts.filter(alert_level=Alert.Level.CRITICAL).count(),
        "security_events": logs.count(),
        "alerts": list(alerts.values("alert_type", "alert_level", "message", "created_at")[:50]),
    }


def generate_csv(data, filepath):
    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["SNMADMDCP Report"])
        for key, val in data.items():
            if key != "alerts":
                writer.writerow([key, val])
        writer.writerow([])
        writer.writerow(["Alerts"])
        writer.writerow(["Type", "Level", "Message", "Created"])
        for a in data.get("alerts", []):
            writer.writerow([a.get("alert_type"), a.get("alert_level"), a.get("message"), a.get("created_at")])


def generate_excel(data, filepath):
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    row = 1
    for key, val in data.items():
        if key != "alerts":
            ws.cell(row=row, column=1, value=key)
            ws.cell(row=row, column=2, value=str(val))
            row += 1
    ws2 = wb.create_sheet("Alerts")
    ws2.append(["Type", "Level", "Message", "Created"])
    for a in data.get("alerts", []):
        ws2.append([a.get("alert_type"), a.get("alert_level"), a.get("message"), str(a.get("created_at"))])
    wb.save(filepath)


def generate_pdf(data, filepath):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    doc = SimpleDocTemplate(filepath, pagesize=letter)
    styles = getSampleStyleSheet()
    story = [Paragraph("SNMADMDCP Network Report", styles["Title"]), Spacer(1, 12)]
    rows = [["Metric", "Value"]]
    for key, val in data.items():
        if key != "alerts":
            rows.append([key, str(val)])
    t = Table(rows, colWidths=[200, 300])
    t.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
            ]
        )
    )
    story.append(t)
    doc.build(story)


def generate_report_file(report):
    period_start, period_end = _period_bounds(report.report_type)
    report.period_start = period_start
    report.period_end = period_end
    data = collect_report_data(period_start, period_end)
    reports_dir = os.path.join(settings.MEDIA_ROOT, "reports")
    os.makedirs(reports_dir, exist_ok=True)
    ext = report.export_format
    filename = f"report_{report.id}_{report.report_type}.{ext if ext != 'excel' else 'xlsx'}"
    filepath = os.path.join(reports_dir, filename)
    if ext == "csv":
        generate_csv(data, str(filepath))
    elif ext == "excel":
        generate_excel(data, str(filepath))
    else:
        generate_pdf(data, str(filepath))
    report.file_path = f"reports/{filename}"
    report.save(update_fields=["file_path", "period_start", "period_end"])
    return report
